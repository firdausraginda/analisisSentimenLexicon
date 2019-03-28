from openpyxl import load_workbook
from lexiconCustom import lexCustom
from nltk.tokenize import sent_tokenize, word_tokenize

dataset = load_workbook('../dataset_TA/AI_EDOM_ganjil_18_19.xlsx', data_only=True)
ai_sjn = dataset['AI-SJN']

def importExcelDataSet():
    hasil = []
    labelManual = []
    for i in range(2, 105):
        if (ai_sjn.cell(row=i, column=7).value == None):
            break
        else:
            hasil.append(ai_sjn.cell(row=i, column=7).value)
            labelManual.append(ai_sjn.cell(row=i, column=11).value)
    return hasil, labelManual

a, b = importExcelDataSet()

print(b)

# for data in lexCustom:
#     print(data[0])

from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory

factory = StopWordRemoverFactory()
stopwords = factory.get_stop_words()
print(stopwords)