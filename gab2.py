# -------------komentar-------------
# *sudah bisa handle ngram = 2 dan 3 namun masih ada masalah
# *belum bisa handle kata yg di pisah dgn strip misal = hati-hati
# *belum looping smua dataset

# -------------register lib-------------
# import excel
from openpyxl import load_workbook
dataset = load_workbook('../dataset_TA/coba.xlsx')
sheet1 = dataset['Sheet1']
dataset = load_workbook('../../inset lexicon/InSet-Lexicon/inset.xlsx')
negatif = dataset['negatif']
positif = dataset['positif']

# tokenization
from nltk.tokenize import sent_tokenize, word_tokenize

# stemming
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
factory = StemmerFactory()
stemmer = factory.create_stemmer()

#stopword removal
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()

# -------------global variable-------------
punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
dataStatis = 'aku mencium telapak kaki ayah, sangat ingin makan bawang agar sehat.'
arrPositif = []
arrNegatif = []
hasilNGram1 = []
hasilNGram2 = []
hasilNGram3 = []

# -------------import excel dataset-------------
def importExcelDataSet():
    hasil = []
    for i in range(2,7):
        hasil.append(sheet1.cell(row=i, column=7).value)
    return hasil

# -------------stopword removal-------------
def stopwordRemoval(data):
    hasil = []
    hasil2 = []
    for kata in data:
        hasil.append(stopword.remove(kata))
    for kata in hasil:
        if kata != '':
            hasil2.append(kata) 
    return hasil2

# -------------punctuation removal dan case conversion-------------
def punctuationRemoval(data):
    hasil = []
    for kata in data:
        if kata not in punctuations:
            hasil.append(kata.lower())
    return hasil

# -------------stemming-------------
def stemmingWord(data):
    hasil = []
    for kata in data:
        hasil.append(stemmer.stem(kata))
    return hasil

# -------------tokenisasi-------------
def tokenization(data):
    hasil = word_tokenize(data)
    return hasil

# -------------join string n-gram all-------------
def nGramAll(hasilPraproses):
    nGram1(hasilPraproses)
    nGram2(hasilPraproses)
    nGram3(hasilPraproses)

# -------------join string n-gram = 1-------------
def nGram1(hasilPraproses):
    global hasilNGram1
    
    hasilNGram1 = hasilPraproses

# -------------join string n-gram = 2-------------
def nGram2(hasilPraproses):
    separator = ' '
    tempKata = ''
    global hasilNGram2

    total = int(len(hasilPraproses))

    for i in range(0, total):
        if (i == total-1):
            return
        else:
            tempKata = hasilPraproses[i] + ' ' + hasilPraproses[i+1]
            hasilNGram2.append(tempKata)

# -------------join string n-gram = 3-------------
def nGram3(hasilPraproses):
    separator = ' '
    tempKata = ''
    global hasilNGram3

    total = int(len(hasilPraproses))

    for j in range(0, total):
        if (j == total-2):
            return
        else:
            tempKata = hasilPraproses[j] + ' ' + hasilPraproses[j+1] + ' ' + hasilPraproses[j+2] 
            hasilNGram3.append(tempKata)

# -------------cari senti word positif-------------
def cariSentiWordPositif(param1, param2, param3):
    arrPositifNGram3 = []
    arrPositifNGram2 = []
    global arrPositif

    # cari sentiword dgn n-gram = 3
    for kata in param3:
        for i in range(2,6611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram3.append([sentiWordPositif,weightPositif])

    # hapus kata ngram3 pada ngram1 supaya tidak double
    if (len(arrPositifNGram3) > 0):
        for kata in arrPositifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                param1.remove(kata)

    # cari sentiword dgn n-gram = 2
    for kata in param2:
        for i in range(2,6611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram2.append([sentiWordPositif,weightPositif])

    # hapus kata ngram2 pada ngram1 supaya tidak double
    if (len(arrPositifNGram2) > 0):
        for kata in arrPositifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                param1.remove(kata)

    # cari sentiword dgn n-gram = 1
    for kata in param1:
        for i in range(2,6611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositif.append([sentiWordPositif,weightPositif])

    for kata in arrPositifNGram2:
        arrPositif.append(kata)
    
    for kata in arrPositifNGram3:
        arrPositif.append(kata)

# -------------cari senti word negatif-------------
def cariSentiWordNegatif(param1, param2, param3):
    arrNegatifNGram3 = []
    arrNegatifNGram2 = []
    global arrNegatif

    # cari sentiword dgn n-gram = 3
    for kata in param3:
        for i in range(2,6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordPositif = (negatif.cell(row=i, column=1).value)
                weightPositif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram3.append([sentiWordPositif,weightPositif])

    # hapus kata ngram3 pada ngram1 supaya tidak double
    if (len(arrNegatifNGram3) > 0):
        for kata in arrNegatifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                param1.remove(kata)

    # cari sentiword dgn n-gram = 2
    for kata in param2:
        for i in range(2,6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordPositif = (negatif.cell(row=i, column=1).value)
                weightPositif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram2.append([sentiWordPositif,weightPositif])

    # hapus kata ngram2 pada ngram1 supaya tidak double
    if (len(arrNegatifNGram2) > 0):
        for kata in arrNegatifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                param1.remove(kata)

    # cari sentiword dgn n-gram = 1
    for kata in param1:
        for i in range(2,6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordPositif = (negatif.cell(row=i, column=1).value)
                weightPositif = (negatif.cell(row=i, column=2).value)
                arrNegatif.append([sentiWordPositif,weightPositif])

    for kata in arrNegatifNGram2:
        arrNegatif.append(kata)
    
    for kata in arrNegatifNGram3:
        arrNegatif.append(kata)

# -------------itung sentiment score-------------
def sentimentScore():
    global arrPositif
    global arrNegatif
    countPositif = 0
    countNegatif = 0
    for arr in arrPositif:
        countPositif = countPositif + arr[1]
    for arr in arrNegatif:
        countNegatif = countNegatif + arr[1]
    print('countNegatif: ', countNegatif)
    print('countPositif: ', countPositif)
    sentimentScore = countNegatif + countPositif
    return sentimentScore

# -------------main program-------------
hasilToken = tokenization(dataStatis)
hasilStem = stemmingWord(hasilToken)
hasilNoPuct = punctuationRemoval(hasilStem)
hasilStopWord = stopwordRemoval(hasilNoPuct)

print('hasil praposes: ', hasilStopWord)

nGramAll(hasilStopWord)

print(hasilNGram1)
print(hasilNGram2)
print(hasilNGram3)

cariSentiWordNegatif(hasilNGram1, hasilNGram2, hasilNGram3)
cariSentiWordPositif(hasilNGram1, hasilNGram2, hasilNGram3)

hasilSentimen = sentimentScore()

print('arrPositif: ', arrPositif)
print('arrNegatif: ', arrNegatif)

print('hasil sentimen score: ', hasilSentimen)