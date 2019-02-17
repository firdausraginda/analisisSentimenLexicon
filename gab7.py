# -------------komentar-------------
# *sudah bisa handle n-gram = 2 dan 3
# *sudah bisa handle n-gram = 2 dan 3 yg bukan kata dasar
# *sudah bisa handle kata yg di pisah dgn strip misal = tercengang-cengang
# *sudah bisa looping smua dataset
# *ada pilihan stopword tapi custom stopword masih harus di analisis lagi

# -------------register lib-------------
# import excel
from openpyxl import load_workbook
dataset = load_workbook('../dataset_TA/coba.xlsx')
sheet1 = dataset['Sheet1']
inSetLexicon = load_workbook('../../inset lexicon/InSet-Lexicon/inset.xlsx')
negatif = inSetLexicon['negatif']
positif = inSetLexicon['positif']

# tokenization
from nltk.tokenize import sent_tokenize, word_tokenize

# stemming
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
factory = StemmerFactory()
stemmer = factory.create_stemmer()

#stopword removal
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()
getStopWord = factory.get_stop_words()

# -------------global variable-------------
dataStatis = 'buang air kecil karena kurang pikiran, semua pergi dilakukan untuk mencari pasangan hidup, serta bersuka-sukaan.'
# dataStatis = 'aku mencium telapak kaki ayah, sangat ingin makan bawang agar sehat.'
# dataStatis = 'aku jalan bebas hambat, tarik napas habis, membuat orak senyum di jalan.'
# dataStatis = 'Jangan terlalu sering memenggal lidah, suka mengganggu kita yang mau belajar suka mencium telapak kaki ayah.'
# dataStatis = 'hati kecil buang air kecil, kecil hati, buang hati air'
# dataStatis = 'menyapu-nyapu lantai dengan sapu, seolah-olah semangatku berkobar-kobar untuk menyapu-nyapu kobar'

# -------------import excel dataset-------------
def importExcelDataSet():
    hasil = []
    labelManual = []
    for i in range(2, 48):
        hasil.append(sheet1.cell(row=i, column=7).value)
        labelManual.append(sheet1.cell(row=i, column=8).value)
    return hasil, labelManual

# -------------stopword removal-------------
def stopwordRemoval(data):
    hasil = []
    hasilBersih = []
    stopwordCustom = []
    # nonStopWord = ['tidak', 'sementara', 'boleh', 'nggak', 'harus', 'amat', 'supaya', 'agar']
    nonStopWord = ['tidak']
    
    # default dari stopword sastrawi
    for kata in data:
        hasil.append(stopword.remove(kata))
    for kata in hasil:
        if kata != '':
            hasilBersih.append(kata) 

    # custom stopword
    # isiStopWord = getStopWord
    # for isi in isiStopWord:
    #     if isi not in nonStopWord:
    #         stopwordCustom.append(isi)
    # for kata in data:
    #     if kata not in stopwordCustom:
    #         hasil.append(kata)

    return hasilBersih

# -------------punctuation removal dan case conversion-------------
def punctuationRemoval(data):
    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    hasil = []
    for kata in data:
        if kata not in punctuations:
            hasil.append(kata.lower())
    return hasil

# -------------stemming-------------
def stemmingWord(data):
    hasil = []
    for kata in data:
        hasil.append(stemmer.stem(kata))
    return hasil

# -------------tokenisasi-------------
def tokenization(data):
    hasil = word_tokenize(data)
    return hasil

# -------------join string n-gram all-------------
def nGramAll(hasilPraproses):
    hasilNGram1 = nGram1(hasilPraproses)
    hasilNGram2 = nGram2(hasilPraproses)
    hasilNGram3 = nGram3(hasilPraproses)
    return hasilNGram1, hasilNGram2, hasilNGram3

# -------------join string n-gram = 1-------------
def nGram1(hasilPraproses):
    return hasilPraproses

# -------------join string n-gram = 2-------------
def nGram2(hasilPraproses):
    tempKata = ''
    hasilNGram2 = []

    total = int(len(hasilPraproses))

    for i in range(0, total):
        if (i == total-1):
            return hasilNGram2
        else:
            tempKata = hasilPraproses[i] + ' ' + hasilPraproses[i+1]
            hasilNGram2.append(tempKata)
    
    return hasilNGram2

# -------------join string n-gram = 3-------------
def nGram3(hasilPraproses):
    tempKata = ''
    hasilNGram3 = []

    total = int(len(hasilPraproses))

    if (total >= 3):
        for j in range(0, total):
            if (j == total-2):
                return hasilNGram3
            else:
                tempKata = hasilPraproses[j] + ' ' + hasilPraproses[j+1] + ' ' + hasilPraproses[j+2] 
                hasilNGram3.append(tempKata)
   
    return hasilNGram3

# -------------cari senti word sebelum stemming-------------
def sentiWordBeforeStem(hasilToken):
    arrPositifNGram3 = []
    arrNegatifNGram3 = []
    arrPositifNGram2 = []
    arrNegatifNGram2 = []
    arrPositifNGram1Strip = []
    arrNegatifNGram1Strip = []
    hasilNGramPositif = []
    hasilNGramNegatif = []

    hasilNGram1, hasilNGram2, hasilNGram3 = nGramAll(hasilToken)

    # cari sentiword dgn n-gram = 3
    for kata in hasilNGram3:
        for i in range(2,3611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram3.append([sentiWordPositif,weightPositif])
        for i in range(2, 6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram3.append([sentiWordNegatif,weightNegatif])
        
    # hapus kata ngram3 pada ngram1 dan ngram2 supaya tidak double
    if (len(arrPositifNGram3) > 0):
        for kata in arrPositifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
                for kataParam in hasilNGram2:
                    if kata in kataParam:
                        hasilNGram2.remove(kataParam)
                        break
                            
    if (len(arrNegatifNGram3) > 0):
        for kata in arrNegatifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
                for kataParam in hasilNGram2:
                    if kata in kataParam:
                        hasilNGram2.remove(kataParam)
                        break

    # cari sentiword dgn n-gram = 2
    for kata in hasilNGram2:
        for i in range(2,3611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram2.append([sentiWordPositif,weightPositif])
        for i in range(2,6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram2.append([sentiWordNegatif,weightNegatif])
    
    # hapus kata ngram2 pada ngram1 supaya tidak double
    if (len(arrPositifNGram2) > 0):
        for kata in arrPositifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
    if (len(arrNegatifNGram2) > 0):
        for kata in arrNegatifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
    
    # cari sentiword dgn n-gram = 1 yang pake strip, misal = 'berkobar-kobar'
    for kata in hasilNGram1:
        if ('-' in kata):
            for i in range(2,3611):
                if (kata == positif.cell(row=i, column=1).value) :
                    sentiWordPositif = (positif.cell(row=i, column=1).value)
                    weightPositif = (positif.cell(row=i, column=2).value)
                    arrPositifNGram1Strip.append([sentiWordPositif,weightPositif])
            for i in range(2,6611):
                if (kata == negatif.cell(row=i, column=1).value) :
                    sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                    weightNegatif = (negatif.cell(row=i, column=2).value)
                    arrNegatifNGram1Strip.append([sentiWordNegatif,weightNegatif])

    # hapus kata ngram1 yg pake strip supaya tidak double
    if (len(arrPositifNGram1Strip) > 0):
        for kata in arrPositifNGram1Strip:
            for kataParam in hasilNGram1:
                if kataParam == kata[0]:
                    hasilNGram1.remove(kataParam)
                    break
    if (len(arrNegatifNGram1Strip) > 0):
        for kata in arrNegatifNGram1Strip:
            for kataParam in hasilNGram1:
                if kataParam == kata[0]:
                    hasilNGram1.remove(kataParam)
                    break

    for kata in arrPositifNGram1Strip:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram1Strip:
        hasilNGramNegatif.append(kata)
    for kata in arrPositifNGram2:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram2:
        hasilNGramNegatif.append(kata)
    for kata in arrPositifNGram3:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram3:
        hasilNGramNegatif.append(kata)
    
    return hasilNGramPositif, hasilNGramNegatif, hasilNGram1

# -------------cari senti word setelah stemming-------------
def sentiWordAfterStem(hasilPraproses, paramNGramPositif, paramNGramNegatif):
    arrPositif = []
    arrNegatif = []

    # cari sentiword dgn n-gram = 1
    for kata in hasilPraproses:
        for i in range(2,6611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositif.append([sentiWordPositif,weightPositif])
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatif.append([sentiWordNegatif,weightNegatif])

    for kata in paramNGramPositif:
        arrPositif.append(kata)
    for kata in paramNGramNegatif:
        arrNegatif.append(kata)
    
    print('arrPositif: ', arrPositif)
    print('arrNegatif: ', arrNegatif)
    
    return arrPositif, arrNegatif

# -------------itung sentiment score-------------
def sentimentScore(hasilPositif, hasilNegatif):
    countPositif = 0
    countNegatif = 0
    
    for arr in hasilPositif:
        countPositif = countPositif + arr[1]
    for arr in hasilNegatif:
        countNegatif = countNegatif + arr[1]

    print('countPositif: ', countPositif)
    print('countNegatif: ', countNegatif)

    sentimentScore = countNegatif + countPositif
 
    return sentimentScore

# -------------cek sentimen-------------
def cekSentimen(nilaiSentimen):
    hasil = ''
    if (nilaiSentimen > 0):
        hasil = 'positif'
    elif (nilaiSentimen < 0):
        hasil = 'negatif'
    elif (nilaiSentimen == 0):
        hasil = 'netral'
    
    return hasil

# -------------looping hasil program-------------
def loopHasilProgram(hasilProgram):
    idx = 0
    for data in hasilProgram:
        idx += 1
        print('hasil ke-', idx, ': ', data)

# -------------itung akurasi sistem-------------
def akurasiSistem():
    jmlData = int(len(hasilImport))
    countTrue = 0
    hasil = 0
    yangSalah = []

    for i in range(0, jmlData):
        if (hasilLabelManual[i] == hasilLoop[i][2]):
            countTrue += 1
        elif (hasilLabelManual[i] != hasilLoop[i][2]):
            yangSalah.append(hasilLoop[i])

    hasil = (countTrue / jmlData)*100
    return hasil, yangSalah

# -------------main program-------------
hasilLoop = []

hasilImport, hasilLabelManual = importExcelDataSet()

for dataDinamis in hasilImport:
    hasilToken = tokenization(dataDinamis)
    ngramPositif, ngramNegatif, ngram1 = sentiWordBeforeStem(hasilToken)
    hasilStem = stemmingWord(ngram1)
    hasilNoPuct = punctuationRemoval(hasilStem)
    hasilStopWord = stopwordRemoval(hasilNoPuct)
    hasilPraprosesCoding = hasilStopWord
    
    print('kalimat lengkap : ', dataDinamis)
    print('hasil praposes: ', hasilPraprosesCoding)

    hasilPositif, hasilNegatif = sentiWordAfterStem(hasilPraprosesCoding, ngramPositif, ngramNegatif)
    hasilSentimen = sentimentScore(hasilPositif, hasilNegatif)
    hasilCekSentimen = cekSentimen(hasilSentimen)
    hasilLoop.append([dataDinamis, hasilSentimen, hasilCekSentimen])

    print('hasil sentimen score: ', hasilSentimen)
    print('-----------------------------')

print('-------------LOOPING HASIL PROGRAM-------------')
loopHasilProgram(hasilLoop)

print('-------------AKURASI SISTEM-------------')
hasilAkurasi, dataSalahKlasifikasi = akurasiSistem()
print('akurasi sistem: ', hasilAkurasi)
print('data yg salah di klasifikasi: ', dataSalahKlasifikasi)