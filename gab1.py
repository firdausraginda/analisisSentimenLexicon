# -------------register lib-------------

# import excel
from openpyxl import load_workbook
dataset = load_workbook('../dataset_TA/coba.xlsx')
sheet1 = dataset['Sheet1']
dataset = load_workbook('../../inset lexicon/InSet-Lexicon/inset.xlsx')
negatif = dataset['negatif']
positif = dataset['positif']

# tokenization
from nltk.tokenize import sent_tokenize, word_tokenize

# stemming
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
factory = StemmerFactory()
stemmer = factory.create_stemmer()

#stopword removal
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()

# -------------global variable-------------
punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
dataStatis = 'Pak andit kalau mengajar bicaranya cepat sekali jadi sulit untuk memahami materi pelajaran dikelas. Lebih baik saat mengajar jangan terlalu cepat karena materinya sulit dan akan lebih sulit untuk dipahami yang jika penjelasannya terlalu cepat.'
arrPositif = []
arrNegatif = []

# -------------itung sentiment score-------------
def sentimentScore():
    global arrPositif
    global arrNegatif
    countPositif = 0
    countNegatif = 0
    for arr in arrPositif:
        countPositif = countPositif + arr[1]
    for arr in arrNegatif:
        countNegatif = countNegatif + arr[1]
    print('countNegatif: ', countNegatif)
    print('countPositif: ', countPositif)
    sentimentScore = countNegatif + countPositif
    return sentimentScore

# -------------cari senti word-------------
def cariSentiWord(kata):
    global arrPositif
    global arrNegatif
    for i in range(2,3611):
        if (kata == positif.cell(row=i, column=1).value) :
            sentiWordPositif = (positif.cell(row=i, column=1).value)
            weightPositif = (positif.cell(row=i, column=2).value)
            arrPositif.append([sentiWordPositif,weightPositif])
    for i in range(2,6611):
        if (kata == negatif.cell(row=i, column=1).value) :
            sentiWordNegatif = (negatif.cell(row=i, column=1).value)
            weightNegatif = (negatif.cell(row=i, column=2).value)
            arrNegatif.append([sentiWordNegatif,weightNegatif])

# -------------import excel dataset-------------
def importExcelDataSet():
    hasil = []
    for i in range(2,7):
        hasil.append(sheet1.cell(row=i, column=7).value)
    return hasil

# -------------stopword removal-------------
def stopwordRemoval(data):
    hasil = []
    hasil2 = []
    for kata in data:
        hasil.append(stopword.remove(kata))
    for kata in hasil:
        if kata != '':
            hasil2.append(kata) 
    return hasil2

# -------------punctuation removal dan case conversion-------------
def punctuationRemoval(data):
    hasil = []
    for kata in data:
        if kata not in punctuations:
            hasil.append(kata.lower())
    return hasil

# -------------stemming-------------
def stemmingWord(data):
    hasil = []
    for kata in data:
        hasil.append(stemmer.stem(kata))
    return hasil

# -------------tokenisasi-------------
def tokenization(data):
    hasil = word_tokenize(data)
    return hasil

# -------------main program-------------
hasilToken = tokenization(dataStatis)
hasilStem = stemmingWord(hasilToken)
hasilNoPuct = punctuationRemoval(hasilStem)
hasilStopWord = stopwordRemoval(hasilNoPuct)

print('hasil praposes: ', hasilStopWord)

for kata in hasilStopWord:
    cariSentiWord(kata)

print('arrPositif: ', arrPositif)
print('arrNegatif: ', arrNegatif)

hasilSentimen = sentimentScore()
print('sentiment score: ', hasilSentimen)

# -------------main program-------------
# hasilLoop = []

# hasilImport = importExcelDataSet()

# for data in hasilImport:
#     hasilToken = tokenization(data)
#     hasilStem = stemmingWord(hasilToken)
#     hasilNoPuct = punctuationRemoval(hasilStem)
#     hasilStopWord = stopwordRemoval(hasilNoPuct)
#     hasilLoop.append(hasilStopWord)
# print(hasilLoop)