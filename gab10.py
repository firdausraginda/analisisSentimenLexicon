# -------------komentar-------------
# *sudah bisa handle n-gram = 2 dan 3
# *sudah bisa handle n-gram = 2 dan 3 yg bukan kata dasar
# *sudah bisa handle kata yg di pisah dgn strip misal = tercengang-cengang
# *sudah bisa looping smua dataset
# *dataset matkul AI semua dosen
# *buat lexicon baru
# *looping semua data AI
# *untuk ngitung akurasi, precision, recall, f-measure lvl kalimat dan dokumen
# *pake custom stopword removal

# -------------register lib-------------
# import excel
import openpyxl
getSheets = openpyxl.load_workbook('../dataset_TA/AI_EDOM_ganjil_18_19_siap_sidang.xlsx')
from openpyxl import load_workbook
dataset = load_workbook('../dataset_TA/AI_EDOM_ganjil_18_19_siap_sidang.xlsx', data_only=True)
eval_all = dataset['eval-all']
dosen_1 = dataset['DOSEN-1']
dosen_2 = dataset['DOSEN-2']
dosen_3 = dataset['DOSEN-3']
dosen_4 = dataset['DOSEN-4']
dosen_5 = dataset['DOSEN-5']
dosen_6 = dataset['DOSEN-6']
dosen_7 = dataset['DOSEN-7']
dosen_8 = dataset['DOSEN-8']
dosen_9 = dataset['DOSEN-9']
dosen_10 = dataset['DOSEN-10']
dosen_11 = dataset['DOSEN-11']
dataAI = [dosen_1, dosen_2, dosen_3, dosen_4, dosen_5, dosen_6, dosen_7, dosen_8, dosen_9, dosen_10, dosen_11]
inSetLexicon = load_workbook('../inset lexicon/InSet-Lexicon/inset.xlsx')
negatif = inSetLexicon['negatif']
positif = inSetLexicon['positif']

# write excel
from openpyxl import Workbook
from openpyxl.styles import Font
book = Workbook()

# tokenization
from nltk.tokenize import sent_tokenize, word_tokenize

# stemming
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
factory = StemmerFactory()
stemmer = factory.create_stemmer()

#stopword removal
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
factory = StopWordRemoverFactory()
stopword = factory.create_stop_word_remover()
getStopWord = factory.get_stop_words()

#import custom lexicon
from lexiconCustom import lexCustom

# -------------import excel dataset-------------
def importExcelDataSet(selectSheet):
    hasil = []
    labelManual = []
    labelManual2 = []

    for i in range(2, 105):
        if (selectSheet.cell(row=i, column=7).value == None):
            break
        else:
            hasil.append(selectSheet.cell(row=i, column=7).value)
            labelManual.append(selectSheet.cell(row=i, column=11).value)
    
    labelManual2 = selectSheet.cell(row=2, column=16).value

    return hasil, labelManual, labelManual2

# -------------write excel-------------
def writeExcel(dataHasilPrediksi, sentimentScoreDok, idx):
    sheet = book.create_sheet(getSheets.sheetnames[idx], -1)
    
    sheet['A1'] = 'Feedback'
    sheet['B1'] = 'Sentiment Score'
    sheet['C1'] = 'Sentiment'
    sheet['E1'] = 'Sentiment Score Dokumen ini'
    sheet['E2'] = sentimentScoreDok

    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)
    sheet['E1'].font = Font(bold=True)

    for i in range(2, len(dataHasilPrediksi)+2):
        sheet.cell(row=i, column=1).value = dataHasilPrediksi[i-2][0]
        sheet.cell(row=i, column=2).value = dataHasilPrediksi[i-2][1]
        sheet.cell(row=i, column=3).value = dataHasilPrediksi[i-2][2]

    book.save('hasil sistem semua dokumen.xlsx')

# -------------stopword removal-------------
def stopwordRemoval(data):
    hasil = []
    # hasil2 = []
    stopwordCustom = []
    moreStopWord = ['pa']
    nonStopWord = ['tidak', 'nggak']

    # for kata in data:
    #     hasil.append(stopword.remove(kata))
    # for kata in hasil:
    #     if kata != '':
    #         hasil2.append(kata) 

    # custom stopword
    isiStopWord = getStopWord
    for isi in isiStopWord:
        if isi not in nonStopWord:
            stopwordCustom.append(isi)
    for isi in moreStopWord:
        stopwordCustom.append(isi)
    for kata in data:
        if kata not in stopwordCustom:
            hasil.append(kata)

    return hasil

# -------------punctuation removal dan case conversion-------------
def punctuationRemoval(data):
    punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    hasil = []
    for kata in data:
        if kata not in punctuations:
            hasil.append(kata)
    return hasil

# -------------stemming-------------
def stemmingWord(data):
    hasil = []
    for kata in data:
        hasil.append(stemmer.stem(kata))
    return hasil

# -------------tokenisasi-------------
def tokenization(data):
    data = str(data)
    hasil = word_tokenize(data.lower())
    return hasil

# -------------join string n-gram all-------------
def nGramAll(hasilPraproses):
    hasilNGram1 = nGram1(hasilPraproses)
    hasilNGram2 = nGram2(hasilPraproses)
    hasilNGram3 = nGram3(hasilPraproses)
    return hasilNGram1, hasilNGram2, hasilNGram3

# -------------join string n-gram = 1-------------
def nGram1(hasilPraproses):
    return hasilPraproses

# -------------join string n-gram = 2-------------
def nGram2(hasilPraproses):
    tempKata = ''
    hasilNGram2 = []

    total = int(len(hasilPraproses))

    for i in range(0, total):
        if (i == total-1):
            return hasilNGram2
        else:
            tempKata = hasilPraproses[i] + ' ' + hasilPraproses[i+1]
            hasilNGram2.append(tempKata)
    
    return hasilNGram2

# -------------join string n-gram = 3-------------
def nGram3(hasilPraproses):
    tempKata = ''
    hasilNGram3 = []

    total = int(len(hasilPraproses))

    if (total >= 3):
        for j in range(0, total):
            if (j == total-2):
                return hasilNGram3
            else:
                tempKata = hasilPraproses[j] + ' ' + hasilPraproses[j+1] + ' ' + hasilPraproses[j+2] 
                hasilNGram3.append(tempKata)
   
    return hasilNGram3

# -------------cari senti word sebelum stemming-------------
def sentiWordBeforeStem(hasilToken):
    arrPositifNGram3 = []
    arrNegatifNGram3 = []
    arrPositifNGram2 = []
    arrNegatifNGram2 = []
    arrPositifNGram1Strip = []
    arrNegatifNGram1Strip = []
    hasilNGramPositif = []
    hasilNGramNegatif = []

    hasilNGram1, hasilNGram2, hasilNGram3 = nGramAll(hasilToken)

    # cari sentiword dgn n-gram = 3
    for kata in hasilNGram3:
        for i in range(2,3611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram3.append([sentiWordPositif,weightPositif])
        for i in range(2, 6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram3.append([sentiWordNegatif,weightNegatif])
        
    # hapus kata ngram3 pada ngram1 dan ngram2 supaya tidak double
    if (len(arrPositifNGram3) > 0):
        for kata in arrPositifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
                for kataParam in hasilNGram2:
                    if kata in kataParam:
                        hasilNGram2.remove(kataParam)
                        break

    if (len(arrNegatifNGram3) > 0):
        for kata in arrNegatifNGram3:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
                for kataParam in hasilNGram2:
                    if kata in kataParam:
                        hasilNGram2.remove(kataParam)
                        break

    # cari sentiword di lexicon custom dgn n-gram = 2
    for kata in hasilNGram2:
        for i in range(0, len(lexCustom)):
            if (kata == lexCustom[i][0]):
                sentiWordPositif = lexCustom[i][0]
                weightPositif = lexCustom[i][1]
                arrPositifNGram2.append([sentiWordPositif,weightPositif])

    # cari sentiword dgn n-gram = 2
    for kata in hasilNGram2:
        for i in range(2,3611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositifNGram2.append([sentiWordPositif,weightPositif])
        for i in range(2,6611):
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatifNGram2.append([sentiWordNegatif,weightNegatif])
    
    # hapus kata ngram2 pada ngram1 supaya tidak double
    if (len(arrPositifNGram2) > 0):
        for kata in arrPositifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
    if (len(arrNegatifNGram2) > 0):
        for kata in arrNegatifNGram2:
            token_temp = (word_tokenize(kata[0]))
            for kata in token_temp:
                for kataParam in hasilNGram1:
                    if kataParam == kata:
                        hasilNGram1.remove(kataParam)
                        break
    
    # cari sentiword di lexicon custom dgn n-gram = 1
    for kata in hasilNGram1:
        for i in range(0, len(lexCustom)):
            if (kata == lexCustom[i][0]):
                sentiWordPositif = lexCustom[i][0]
                weightPositif = lexCustom[i][1]
                arrPositifNGram1Strip.append([sentiWordPositif,weightPositif])

    # cari sentiword dgn n-gram = 1 yang pake strip, misal = 'berkobar-kobar'
    for kata in hasilNGram1:
        if ('-' in kata):
            for i in range(2,3611):
                if (kata == positif.cell(row=i, column=1).value) :
                    sentiWordPositif = (positif.cell(row=i, column=1).value)
                    weightPositif = (positif.cell(row=i, column=2).value)
                    arrPositifNGram1Strip.append([sentiWordPositif,weightPositif])
            for i in range(2,6611):
                if (kata == negatif.cell(row=i, column=1).value) :
                    sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                    weightNegatif = (negatif.cell(row=i, column=2).value)
                    arrNegatifNGram1Strip.append([sentiWordNegatif,weightNegatif])

    # hapus kata ngram1 yg pake strip supaya tidak double
    if (len(arrPositifNGram1Strip) > 0):
        for kata in arrPositifNGram1Strip:
            for kataParam in hasilNGram1:
                if kataParam == kata[0]:
                    hasilNGram1.remove(kataParam)
                    break
    if (len(arrNegatifNGram1Strip) > 0):
        for kata in arrNegatifNGram1Strip:
            for kataParam in hasilNGram1:
                if kataParam == kata[0]:
                    hasilNGram1.remove(kataParam)
                    break

    for kata in arrPositifNGram1Strip:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram1Strip:
        hasilNGramNegatif.append(kata)
    for kata in arrPositifNGram2:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram2:
        hasilNGramNegatif.append(kata)
    for kata in arrPositifNGram3:
        hasilNGramPositif.append(kata)
    for kata in arrNegatifNGram3:
        hasilNGramNegatif.append(kata)
    
    return hasilNGramPositif, hasilNGramNegatif, hasilNGram1

# -------------cari senti word setelah stemming-------------
def sentiWordAfterStem(hasilPraproses, paramNGramPositif, paramNGramNegatif):
    arrPositif = []
    arrNegatif = []

    # cari sentiword dgn n-gram = 1
    for kata in hasilPraproses:
        for i in range(2,6611):
            if (kata == positif.cell(row=i, column=1).value) :
                sentiWordPositif = (positif.cell(row=i, column=1).value)
                weightPositif = (positif.cell(row=i, column=2).value)
                arrPositif.append([sentiWordPositif,weightPositif])
            if (kata == negatif.cell(row=i, column=1).value) :
                sentiWordNegatif = (negatif.cell(row=i, column=1).value)
                weightNegatif = (negatif.cell(row=i, column=2).value)
                arrNegatif.append([sentiWordNegatif,weightNegatif])

    for kata in paramNGramPositif:
        arrPositif.append(kata)
    for kata in paramNGramNegatif:
        arrNegatif.append(kata)
    
    # print('arrPositif: ', arrPositif)
    # print('arrNegatif: ', arrNegatif)
    
    return arrPositif, arrNegatif

# -------------itung sentiment score-------------
def sentimentScore(hasilPositif, hasilNegatif):
    countPositif = 0
    countNegatif = 0
    
    for arr in hasilPositif:
        countPositif = countPositif + arr[1]
    for arr in hasilNegatif:
        countNegatif = countNegatif + arr[1]

    # print('countPositif: ', countPositif)
    # print('countNegatif: ', countNegatif)

    sentimentScore = countNegatif + countPositif
 
    return sentimentScore

# -------------cek sentimen-------------
def cekSentimen(nilaiSentimen):
    hasil = ''
    if (nilaiSentimen > 0):
        hasil = 'positif'
    elif (nilaiSentimen < 0):
        hasil = 'negatif'
    elif (nilaiSentimen == 0):
        hasil = 'netral'
    
    return hasil

# -------------looping hasil program-------------
def loopHasilProgram(hasilProgram):
    idx = 0
    for data in hasilProgram:
        idx += 1
        print('hasil ke-', idx, ': ', data)

# -------------itung precision dan recall-------------
def hitungPreRec(truePositif, pembagi1, pembagi2, pembagi3):
    hasilPrecRec = None
    pembagi = pembagi1 + pembagi2 + pembagi3
    if (pembagi == 0) :
        hasilPrecRec = 'none'
    else:
        hasilPrecRec = truePositif / pembagi 
    
    return hasilPrecRec

# -------------itung F-Measure-------------
def itungFMeasure(pre, rec, beta):
    hasil = None
    if (pre == 'none' or rec == 'none'):
            hasil = 'none'
    else:
        tempPembagi = (beta*beta*pre)+rec
        if (tempPembagi == 0):
            hasil = 'none'
        else:
            hasil = (((beta*beta) + 1)*pre*rec)/((beta*beta*pre)+rec)
    return hasil

# -------------itung evaluasi sistem-------------
def evaluasiSistem(labelManualParam, loopSistem):
    jmlData = int(len(loopSistem))
    countNilaiTotal = 0
    arraySalah = []

    countPosPos = 0
    countPosNeg = 0
    countPosNet = 0
    countNegPos = 0
    countNegNeg = 0
    countNegNet = 0
    countNetPos = 0
    countNetNeg = 0
    countNetNet = 0

    for i in range(0, jmlData):
        # positif
        if (labelManualParam[i] == 'positif' and loopSistem[i][2] == 'positif'):
            countPosPos += 1
        elif(labelManualParam[i] == 'positif' and loopSistem[i][2] == 'negatif'):
            countPosNeg += 1
            arraySalah.append(loopSistem[i])
        elif(labelManualParam[i] == 'positif' and loopSistem[i][2] == 'netral'):
            countPosNet += 1
            arraySalah.append(loopSistem[i])
        # negatif
        elif(labelManualParam[i] == 'negatif' and loopSistem[i][2] == 'positif'):
            countNegPos += 1
            arraySalah.append(loopSistem[i])
        elif(labelManualParam[i] == 'negatif' and loopSistem[i][2] == 'negatif'):
            countNegNeg += 1
        elif(labelManualParam[i] == 'negatif' and loopSistem[i][2] == 'netral'):
            countNegNet += 1
            arraySalah.append(loopSistem[i])
        # netral
        elif(labelManualParam[i] == 'netral' and loopSistem[i][2] == 'positif'):
            countNetPos += 1
            arraySalah.append(loopSistem[i])
        elif(labelManualParam[i] == 'netral' and loopSistem[i][2] == 'negatif'):
            countNetNeg += 1
            arraySalah.append(loopSistem[i])
        elif(labelManualParam[i] == 'netral' and loopSistem[i][2] == 'netral'):
            countNetNet += 1
        # total sentiment score
        countNilaiTotal += loopSistem[i][1]

    # precision positif
    prePos = hitungPreRec(countPosPos, countPosPos, countNegPos, countNetPos)
    # precision negatif
    preNeg = hitungPreRec(countNegNeg, countPosNeg, countNegNeg, countNetNeg)
    # precision netral        
    preNet = hitungPreRec(countNetNet, countPosNet, countNegNet, countNetNet)

    # recall positif
    recPos = hitungPreRec(countPosPos, countPosPos, countPosNeg, countPosNet)
    # recall negatif
    recNeg = hitungPreRec(countNegNeg, countNegPos, countNegNeg, countNegNet)
    # recall netral
    recNet = hitungPreRec(countNetNet, countNetPos, countNetNeg, countNetNet)

    accuracy = ((countPosPos + countNegNeg + countNetNet) / (countPosPos + countNegPos + countNetPos + countPosNeg + countNegNeg + countNetNeg + countPosNet + countNegNet + countNetNet)) * 100
    confusionMatrix = countPosPos, countNegPos, countNetPos, countPosNeg, countNegNeg, countNetNeg, countPosNet, countNegNet, countNetNet
    precision = prePos, preNeg, preNet
    recall = recPos, recNeg, recNet
    
    # f-measure positif
    fmPos = itungFMeasure(prePos, recPos, 1)
    # f-measure negatif
    fmNeg = itungFMeasure(preNeg, recNeg, 1)
    # f-measure netral
    fmNet = itungFMeasure(preNet, recNet, 1)
    # f-measure semua
    fMeasure = fmPos, fmNeg, fmNet

    return countNilaiTotal, confusionMatrix, accuracy, precision, recall, arraySalah, fMeasure

# -------------main program-------------
hasilLoop2 = []
labelManualArr = []
iterationDokumen = 0

for loopData in dataAI:
    
    hasilLoop = []
    
    hasilImport, hasilLabelManual, hasilLabelManual2 = importExcelDataSet(loopData)
    labelManualArr.append(hasilLabelManual2)
    
    for dataDinamis in hasilImport:
        hasilToken = tokenization(dataDinamis)
        ngramPositif, ngramNegatif, ngram1 = sentiWordBeforeStem(hasilToken)
        hasilStem = stemmingWord(ngram1)
        hasilNoPuct = punctuationRemoval(hasilStem)
        hasilStopWord = stopwordRemoval(hasilNoPuct)
        hasilPraprosesCoding = hasilStopWord
        
        # print('kalimat lengkap : ', dataDinamis)
        # print('hasil praposes: ', hasilPraprosesCoding)

        hasilPositif, hasilNegatif = sentiWordAfterStem(hasilPraprosesCoding, ngramPositif, ngramNegatif)
        hasilSentimen = sentimentScore(hasilPositif, hasilNegatif)
        hasilCekSentimen = cekSentimen(hasilSentimen)
        hasilLoop.append([dataDinamis, hasilSentimen, hasilCekSentimen])

        # print('hasil sentimen score: ', hasilSentimen)
        # print('-----------------------------')

    # print('-------------LOOPING HASIL PROGRAM-------------')
    # loopHasilProgram(hasilLoop)

    print('-------------AKURASI SISTEM LEVEL KALIMAT-------------')
    hasilTotal, conMat, acc, pre, rec, hasilSalah, fMeasure = evaluasiSistem(hasilLabelManual, hasilLoop)
    print('sentiment score total lvl kalimat: ', hasilTotal)
    print('accuracy lvl kalimat: ', acc)
    print('confusion matrix lvl kalimat: ', conMat)
    print('precision lvl kalimat: ', pre)
    print('recall lvl kalimat: ', rec)
    print('F-Measure lvl kalimat: ', fMeasure)

    writeExcel(hasilLoop, hasilTotal, iterationDokumen)
    iterationDokumen += 1

    hasilCekSentimen2 = cekSentimen(hasilTotal)
    hasilLoop2.append([None, hasilTotal, hasilCekSentimen2])

    # print('-------------HASIL YANG SALAH-------------')
    # loopHasilProgram(hasilSalah)

print('---------------------------------------')
print('-------------AKURASI SISTEM LEVEL DOKUMEN-------------')
hasilTotal2, conMat2, acc2, pre2, rec2, hasilSalah2, fMeasure2 = evaluasiSistem(labelManualArr, hasilLoop2)
print('sentiment score total lvl dokumen: ', hasilTotal2)
print('accuracy lvl dokumen: ', acc2)
print('confusion matrix lvl dokumen: ', conMat2)
print('precision lvl dokumen: ', pre2)
print('recall lvl dokumen: ', rec2)
print('F-Measure lvl dokumen: ', fMeasure2)