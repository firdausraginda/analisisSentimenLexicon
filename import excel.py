from openpyxl import load_workbook

# inisialisasi excel untuk read data
wb = load_workbook('./dataset_TA/coba.xlsx')

# get nama sheet dari excel yang di read
sheet1 = wb['Sheet1']

data = []

for i in range(2,7):
    data.append(sheet1.cell(row=i, column=7).value)

# print(data)

