# from openpyxl import load_workbook
# inSetLexicon = load_workbook('../../inset lexicon/InSet-Lexicon/inset.xlsx')
# negatif = inSetLexicon['negatif']
# positif = inSetLexicon['positif']

# for i in range(2, 3611):
#         if (positif.cell(row=i, column=1).value == 'ga'):
#                 print('ga: ', i)
#         if (positif.cell(row=i, column=1).value == 'gak'):
#                 print('gak: ', i)
#         if (positif.cell(row=i, column=1).value == 'nggak'):
#                 print('nggak: ', i)

# =======================================================================================================

from openpyxl import Workbook

book = Workbook()
# sheet = book.active

ws1 = book.create_sheet("lalalala", -1)
ws1['A7'] = 1
ws1.cell(row=2, column=2).value = 2

ws1 = book.create_sheet("xaxaxa", -1)
ws1['A7'] = 2
ws1.cell(row=2, column=2).value = 3

book.save('coba coba.xlsx')